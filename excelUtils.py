import openpyxl
import openpyxl.workbook.workbook as Workbook

def openExcelWorkBook(filePath): # returns xls file
    return openpyxl.load_workbook(filename=filePath)

def openSheet(workbook, sheetName = None):
    return workbook.get_sheet_by_name(sheetName) if sheetName is not None else workbook.active


def getRowCount(filePath):
    workbook = openExcelWorkBook(filePath)
    return openSheet(workbook).max_row

def getColumnCount(filePath):
    workbook = openExcelWorkBook(filePath)
    return openSheet(workbook).max_column

def readExcelDataInRange(filePath, beginningRow = 0, beginningCol = 0):
    workbook = openExcelWorkBook(filePath)
    sheet = openSheet(workbook)
    for r in range(beginningRow, sheet.max_row):
        for c in range(beginningCol, sheet.max_column):
            print(sheet.cell(r,c).value)

def getCellValue(filePath, column, row):
    workbook = openExcelWorkBook(filePath)
    sheet = workbook.get_sheet_by_name()
    return sheet.cell(row, column).value


def writeData(filePath, column, row, data):
    workbook = openExcelWorkBook(filePath)
    sheet = openSheet(workbook)
    sheet.cell(column, row).value = data

