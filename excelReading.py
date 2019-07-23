import openpyxl

pathToExcelFile = "C:\\users\\aahme\\Desktop\\excelTestFile"
excelFile = openpyxl.load_workbook(pathToExcelFile)

sheet = excelFile.active # or we can get sheet by name

for row in range(1,sheet.max_row+1): # doesnt include headers loops through first row, second row.
    for col in range (0,sheet.max_column+1):
        print(sheet.cell(row, col).value) # this gives you the value of the cell



